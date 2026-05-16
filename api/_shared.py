import base64
import hashlib
import hmac
import json
import os
import re
import time
from datetime import datetime, timezone
from io import BytesIO
from urllib.parse import parse_qs, urlencode, urlparse

import openpyxl
import requests
from openai import OpenAI


ROOT = os.path.dirname(os.path.dirname(__file__))
EXCEL_TEMPLATE = os.path.join(ROOT, "Namecard templat for Hubspot.xlsx")

DRIVE_SCOPE = "https://www.googleapis.com/auth/drive.readonly"
USERINFO_SCOPE = "https://www.googleapis.com/auth/userinfo.email"
OPENID_SCOPE = "openid"
SUPPORTED_IMAGE_MIME_PREFIX = "image/"

EXTRACTION_PROMPT = """
You are extracting contact information from a business name card image.

Return a JSON object with exactly these keys:

CORE FIELDS (map to HubSpot template — use null if absent or unreadable):
- first_name
- last_name
- email           (primary email only)
- phone_number    (primary phone; include country code if visible)
- job_title
- company_name
- company_street_address  (full address if present)

SUPPLEMENTARY:
- extra_fields: object with any other info found on the card.
  Common keys: website, linkedin, wechat, fax, secondary_email,
  secondary_phone, department, country, city, postal_code,
  social_media, tagline, industry, registration_number, etc.
  Use snake_case keys. Include everything that appears on the card
  that does not fit the core fields above.

- unclear_fields: array of core field names (from the list above) that
  could not be clearly read. Set those core fields to null.

- raw_text: all visible text on the card exactly as it appears,
  preserving line breaks with \\n.

- notes: brief observations, e.g. "card is in Chinese", "bilingual card",
  "QR code present", "logo only — no address", "handwritten annotation".

RULES:
- Do NOT guess. If unsure, set to null and add to unclear_fields.
- Return ONLY valid JSON. No markdown fences, no explanation.
""".strip()


def env(name, default=None):
    value = os.getenv(name, default)
    if value is None or value == "":
        raise RuntimeError(f"Missing required environment variable: {name}")
    return value


def optional_env(name, default=None):
    value = os.getenv(name)
    return default if value in (None, "") else value


def read_json_body(request):
    length = int(request.headers.get("content-length", "0") or "0")
    if length == 0:
        return {}
    return json.loads(request.rfile.read(length).decode("utf-8"))


def send_json(request, payload, status=200):
    body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    request.send_response(status)
    request.send_header("Content-Type", "application/json; charset=utf-8")
    request.send_header("Content-Length", str(len(body)))
    request.end_headers()
    request.wfile.write(body)


def send_html(request, html, status=200):
    body = html.encode("utf-8")
    request.send_response(status)
    request.send_header("Content-Type", "text/html; charset=utf-8")
    request.send_header("Content-Length", str(len(body)))
    request.end_headers()
    request.wfile.write(body)


def send_bytes(request, data, filename, content_type):
    request.send_response(200)
    request.send_header("Content-Type", content_type)
    request.send_header("Content-Disposition", f'attachment; filename="{filename}"')
    request.send_header("Content-Length", str(len(data)))
    request.end_headers()
    request.wfile.write(data)


def query_params(path):
    return {key: values[-1] for key, values in parse_qs(urlparse(path).query).items()}


def sign_state(workspace_id):
    secret = env("APP_SECRET").encode("utf-8")
    digest = hmac.new(secret, workspace_id.encode("utf-8"), hashlib.sha256).hexdigest()
    return f"{workspace_id}.{digest}"


def verify_state(state):
    if not state or "." not in state:
        raise RuntimeError("Invalid OAuth state.")
    workspace_id, signature = state.rsplit(".", 1)
    expected = sign_state(workspace_id).rsplit(".", 1)[1]
    if not hmac.compare_digest(signature, expected):
        raise RuntimeError("Invalid OAuth state signature.")
    return workspace_id


def oauth_redirect_uri(request=None):
    configured = optional_env("GOOGLE_REDIRECT_URI")
    if configured:
        return configured
    if request:
        host = request.headers.get("host")
        proto = request.headers.get("x-forwarded-proto", "https")
        if host:
            return f"{proto}://{host}/api/oauth_callback"
    raise RuntimeError("Missing GOOGLE_REDIRECT_URI.")


def supabase_headers(prefer=None):
    key = env("SUPABASE_SERVICE_ROLE_KEY")
    headers = {
        "apikey": key,
        "Authorization": f"Bearer {key}",
        "Content-Type": "application/json",
    }
    if prefer:
        headers["Prefer"] = prefer
    return headers


def supabase_url(path):
    return f"{env('SUPABASE_URL').rstrip('/')}/rest/v1/{path.lstrip('/')}"


def supabase_request(method, path, *, params=None, json_body=None, prefer=None):
    response = requests.request(
        method,
        supabase_url(path),
        headers=supabase_headers(prefer=prefer),
        params=params,
        json=json_body,
        timeout=30,
    )
    if response.status_code >= 400:
        raise RuntimeError(f"Supabase error {response.status_code}: {response.text}")
    if not response.text:
        return None
    return response.json()


def upsert_drive_connection(workspace_id, google_email, refresh_token, scopes):
    record = {
        "workspace_id": workspace_id,
        "google_email": google_email,
        "refresh_token": refresh_token,
        "scopes": scopes,
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    supabase_request(
        "POST",
        "drive_connections",
        json_body=record,
        prefer="resolution=merge-duplicates,return=minimal",
    )


def get_drive_connection(workspace_id):
    rows = supabase_request(
        "GET",
        "drive_connections",
        params={
            "workspace_id": f"eq.{workspace_id}",
            "select": "workspace_id,google_email,refresh_token,updated_at",
            "limit": "1",
        },
    )
    return rows[0] if rows else None


def refresh_google_access_token(refresh_token):
    response = requests.post(
        "https://oauth2.googleapis.com/token",
        data={
            "client_id": env("GOOGLE_CLIENT_ID"),
            "client_secret": env("GOOGLE_CLIENT_SECRET"),
            "refresh_token": refresh_token,
            "grant_type": "refresh_token",
        },
        timeout=30,
    )
    if response.status_code >= 400:
        raise RuntimeError(f"Google token refresh failed: {response.text}")
    return response.json()["access_token"]


def google_headers(access_token):
    return {"Authorization": f"Bearer {access_token}"}


def parse_drive_folder_id(folder_input):
    value = (folder_input or "").strip()
    if not value:
        raise RuntimeError("Google Drive folder link or ID is required.")

    patterns = [
        r"/folders/([a-zA-Z0-9_-]+)",
        r"[?&]id=([a-zA-Z0-9_-]+)",
        r"^([a-zA-Z0-9_-]{10,})$",
    ]
    for pattern in patterns:
        match = re.search(pattern, value)
        if match:
            return match.group(1)
    raise RuntimeError("Could not find a folder ID in that Google Drive link.")


def get_drive_file(access_token, file_id, fields):
    response = requests.get(
        f"https://www.googleapis.com/drive/v3/files/{file_id}",
        headers=google_headers(access_token),
        params={"fields": fields, "supportsAllDrives": "true"},
        timeout=30,
    )
    if response.status_code >= 400:
        raise RuntimeError(f"Google Drive metadata error: {response.text}")
    return response.json()


def list_drive_images(access_token, folder_id):
    files = []
    page_token = None
    while True:
        params = {
            "q": f"'{folder_id}' in parents and trashed = false and mimeType contains '{SUPPORTED_IMAGE_MIME_PREFIX}'",
            "fields": "nextPageToken,files(id,name,mimeType,modifiedTime,md5Checksum,size,webViewLink)",
            "pageSize": "100",
            "orderBy": "name",
            "supportsAllDrives": "true",
            "includeItemsFromAllDrives": "true",
        }
        if page_token:
            params["pageToken"] = page_token
        response = requests.get(
            "https://www.googleapis.com/drive/v3/files",
            headers=google_headers(access_token),
            params=params,
            timeout=30,
        )
        if response.status_code >= 400:
            raise RuntimeError(f"Google Drive list error: {response.text}")
        payload = response.json()
        files.extend(payload.get("files", []))
        page_token = payload.get("nextPageToken")
        if not page_token:
            return files


def download_drive_file(access_token, file_id):
    response = requests.get(
        f"https://www.googleapis.com/drive/v3/files/{file_id}",
        headers=google_headers(access_token),
        params={"alt": "media", "supportsAllDrives": "true"},
        timeout=60,
    )
    if response.status_code >= 400:
        raise RuntimeError(f"Google Drive download error: {response.text}")
    return response.content


def source_fingerprint(file_info):
    raw = ":".join(
        [
            file_info.get("id", ""),
            file_info.get("md5Checksum") or file_info.get("modifiedTime") or "",
            file_info.get("size") or "",
        ]
    )
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def find_existing_scan(workspace_id, fingerprint):
    rows = supabase_request(
        "GET",
        "namecard_scans",
        params={
            "workspace_id": f"eq.{workspace_id}",
            "source_fingerprint": f"eq.{fingerprint}",
            "select": "*",
            "limit": "1",
        },
    )
    return rows[0] if rows else None


def insert_scan(workspace_id, folder_id, folder_name, file_info, fingerprint, extraction):
    record = {
        "workspace_id": workspace_id,
        "drive_folder_id": folder_id,
        "drive_folder_name": folder_name,
        "drive_file_id": file_info["id"],
        "drive_file_name": file_info["name"],
        "drive_file_md5": file_info.get("md5Checksum"),
        "drive_modified_time": file_info.get("modifiedTime"),
        "drive_mime_type": file_info.get("mimeType"),
        "source_fingerprint": fingerprint,
        "processed_at": datetime.now(timezone.utc).isoformat(),
        "extraction": extraction,
    }
    rows = supabase_request(
        "POST",
        "namecard_scans",
        params={"on_conflict": "workspace_id,drive_file_id"},
        json_body=record,
        prefer="resolution=merge-duplicates,return=representation",
    )
    return rows[0]


def list_scans(workspace_id):
    return supabase_request(
        "GET",
        "namecard_scans",
        params={
            "workspace_id": f"eq.{workspace_id}",
            "select": "*",
            "order": "processed_at.asc",
        },
    )


def extract_card_from_bytes(image_bytes, mime_type):
    b64 = base64.b64encode(image_bytes).decode("utf-8")
    client = OpenAI(api_key=env("OPENAI_API_KEY"))
    response = client.chat.completions.create(
        model=optional_env("OPENAI_MODEL", "gpt-4o"),
        messages=[
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": EXTRACTION_PROMPT},
                    {
                        "type": "image_url",
                        "image_url": {
                            "url": f"data:{mime_type};base64,{b64}",
                            "detail": "high",
                        },
                    },
                ],
            }
        ],
        max_tokens=1200,
        response_format={"type": "json_object"},
    )
    return json.loads(response.choices[0].message.content)


def build_hubspot_workbook(records):
    wb = openpyxl.load_workbook(EXCEL_TEMPLATE)
    ws = wb.active

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.value = None

    for index, record in enumerate(records, start=2):
        extraction = record["extraction"]
        ws.cell(row=index, column=1, value=extraction.get("first_name"))
        ws.cell(row=index, column=2, value=extraction.get("last_name"))
        ws.cell(row=index, column=3, value=extraction.get("email"))
        ws.cell(row=index, column=4, value=extraction.get("phone_number"))
        ws.cell(row=index, column=5, value=extraction.get("job_title"))
        ws.cell(row=index, column=6, value=extraction.get("company_name"))
        ws.cell(row=index, column=7, value=extraction.get("company_street_address"))

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def google_auth_url(request, workspace_id):
    state = sign_state(workspace_id)
    params = {
        "client_id": env("GOOGLE_CLIENT_ID"),
        "redirect_uri": oauth_redirect_uri(request),
        "response_type": "code",
        "scope": " ".join([OPENID_SCOPE, USERINFO_SCOPE, DRIVE_SCOPE]),
        "access_type": "offline",
        "prompt": "consent",
        "include_granted_scopes": "true",
        "state": state,
    }
    return f"https://accounts.google.com/o/oauth2/v2/auth?{urlencode(params)}"


def exchange_oauth_code(request, code):
    response = requests.post(
        "https://oauth2.googleapis.com/token",
        data={
            "code": code,
            "client_id": env("GOOGLE_CLIENT_ID"),
            "client_secret": env("GOOGLE_CLIENT_SECRET"),
            "redirect_uri": oauth_redirect_uri(request),
            "grant_type": "authorization_code",
        },
        timeout=30,
    )
    if response.status_code >= 400:
        raise RuntimeError(f"Google OAuth exchange failed: {response.text}")
    return response.json()


def fetch_google_email(access_token):
    response = requests.get(
        "https://www.googleapis.com/oauth2/v2/userinfo",
        headers=google_headers(access_token),
        timeout=30,
    )
    if response.status_code >= 400:
        return None
    return response.json().get("email")


def unix_filename_stamp():
    return time.strftime("%Y%m%d-%H%M%S", time.gmtime())
