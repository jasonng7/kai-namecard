#!/usr/bin/env python3
"""
Name card extraction pipeline.

Scans Namecards/ for images, uses OpenAI GPT-4o vision to extract contact info,
writes HubSpot-ready data to the Excel template, and maintains a JSON reference
database with full extraction details for querying and verification.

Usage:
    python3 extract_namecards.py
"""

import base64
import json
import os
import sys
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from dotenv import load_dotenv
from openai import OpenAI

# ── Paths ──────────────────────────────────────────────────────────────────────
ROOT = Path(__file__).parent
NAMECARDS_DIR = ROOT / "Namecards"
EXCEL_TEMPLATE = ROOT / "Namecard templat for Hubspot.xlsx"
DATABASE_FILE = ROOT / "namecard_database.json"
ENV_FILE = ROOT / ".env"

SUPPORTED_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".gif"}

EXTRACTION_PROMPT = """
You are extracting contact information from a business name card image.

Return a JSON object with exactly these keys:

CORE FIELDS (map to HubSpot template — use null if absent or unreadable):
- first_name
- last_name
- email           (primary email only)
- phone_number    (primary phone; include country code if visible)
- job_title
- company_name
- company_street_address  (full address if present)

SUPPLEMENTARY:
- extra_fields: object with any other info found on the card.
  Common keys: website, linkedin, wechat, fax, secondary_email,
  secondary_phone, department, country, city, postal_code,
  social_media, tagline, industry, registration_number, etc.
  Use snake_case keys. Include everything that appears on the card
  that does not fit the core fields above.

- unclear_fields: array of core field names (from the list above) that
  could not be clearly read. Set those core fields to null.

- raw_text: all visible text on the card exactly as it appears,
  preserving line breaks with \\n.

- notes: brief observations, e.g. "card is in Chinese", "bilingual card",
  "QR code present", "logo only — no address", "handwritten annotation".

RULES:
- Do NOT guess. If unsure, set to null and add to unclear_fields.
- Return ONLY valid JSON. No markdown fences, no explanation.
""".strip()


def encode_image(image_path: Path) -> tuple[str, str]:
    mime_map = {
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".png": "image/png",
        ".webp": "image/webp",
        ".gif": "image/gif",
    }
    mime = mime_map.get(image_path.suffix.lower(), "image/jpeg")
    with open(image_path, "rb") as f:
        b64 = base64.b64encode(f.read()).decode("utf-8")
    return b64, mime


def extract_card(client: OpenAI, image_path: Path) -> dict:
    b64, mime = encode_image(image_path)
    response = client.chat.completions.create(
        model="gpt-4o",
        messages=[
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": EXTRACTION_PROMPT},
                    {
                        "type": "image_url",
                        "image_url": {
                            "url": f"data:{mime};base64,{b64}",
                            "detail": "high",
                        },
                    },
                ],
            }
        ],
        max_tokens=1200,
        response_format={"type": "json_object"},
    )
    return json.loads(response.choices[0].message.content)


def load_database() -> list:
    if DATABASE_FILE.exists():
        with open(DATABASE_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return []


def save_database(records: list) -> None:
    with open(DATABASE_FILE, "w", encoding="utf-8") as f:
        json.dump(records, f, indent=2, ensure_ascii=False)
    print(f"  Database saved → {DATABASE_FILE.name}")


def already_processed(records: list, filename: str) -> bool:
    return any(r["source_image"] == filename for r in records)


def update_excel(records: list) -> None:
    wb = openpyxl.load_workbook(EXCEL_TEMPLATE)
    ws = wb.active

    # Clear existing data rows, keep header row 1
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.value = None

    for i, record in enumerate(records, start=2):
        ex = record["extraction"]
        ws.cell(row=i, column=1, value=ex.get("first_name"))
        ws.cell(row=i, column=2, value=ex.get("last_name"))
        ws.cell(row=i, column=3, value=ex.get("email"))
        ws.cell(row=i, column=4, value=ex.get("phone_number"))
        ws.cell(row=i, column=5, value=ex.get("job_title"))
        ws.cell(row=i, column=6, value=ex.get("company_name"))
        ws.cell(row=i, column=7, value=ex.get("company_street_address"))

    wb.save(EXCEL_TEMPLATE)
    print(f"  Excel updated  → {EXCEL_TEMPLATE.name}")


def summarise(extraction: dict) -> str:
    parts = []
    unclear = extraction.get("unclear_fields") or []
    extra = extraction.get("extra_fields") or {}
    if unclear:
        parts.append(f"unclear: {', '.join(unclear)}")
    if extra:
        parts.append(f"extra: {', '.join(extra.keys())}")
    return f" [{'; '.join(parts)}]" if parts else ""


def main() -> None:
    load_dotenv(ENV_FILE)
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        print("ERROR: OPENAI_API_KEY not set in .env file.", file=sys.stderr)
        sys.exit(1)

    client = OpenAI(api_key=api_key)

    images = sorted(
        p for p in NAMECARDS_DIR.iterdir()
        if p.suffix.lower() in SUPPORTED_EXTENSIONS
    )

    if not images:
        print(f"No images found in {NAMECARDS_DIR}/")
        sys.exit(0)

    records = load_database()
    new_count = 0

    print(f"Found {len(images)} image(s) in Namecards/\n")

    for image_path in images:
        if already_processed(records, image_path.name):
            print(f"  [skip]    {image_path.name}  (already in database)")
            continue

        print(f"  [extract] {image_path.name} ...", end=" ", flush=True)
        try:
            extraction = extract_card(client, image_path)
        except Exception as exc:
            print(f"FAILED — {exc}")
            continue

        record = {
            "source_image": image_path.name,
            "processed_at": datetime.now(timezone.utc).isoformat(),
            "extraction": extraction,
        }
        records.append(record)
        save_database(records)
        new_count += 1
        print(f"OK{summarise(extraction)}")

    print()
    if new_count > 0:
        update_excel(records)
        print(f"\n{new_count} new card(s) processed.")
    else:
        print("No new cards to process. Excel unchanged.")

    print(f"\nFiles:\n  {EXCEL_TEMPLATE.name}\n  {DATABASE_FILE.name}")


if __name__ == "__main__":
    main()
